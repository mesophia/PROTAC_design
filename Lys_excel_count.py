import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import random

# Load and filter original data
input_file = "modified_pdbs/Lys_distance_RMSD.xlsx"
df = pd.read_excel(input_file)

pdb_column = df.columns[0]  # PDB file column
lys_columns = df.columns[2:]  # Lysine distance columns
threshold = 30


lys_filtered_data = {}

for lys in lys_columns:
    filtered_pdbs = df[pdb_column][df[lys] <= threshold].dropna().tolist()
    if filtered_pdbs:
        lys_filtered_data[lys] = [len(filtered_pdbs)] + filtered_pdbs


max_len = max(len(v) for v in lys_filtered_data.values())
for k in lys_filtered_data:
    lys_filtered_data[k] += [None] * (max_len - len(lys_filtered_data[k]))

df_filtered = pd.DataFrame(lys_filtered_data)

# Save to Excel
filtered_output_file = "modified_pdbs/Lys_filtered_RMSD.xlsx"
df_filtered.to_excel(filtered_output_file, index=False)


wb = load_workbook(filtered_output_file)
ws = wb.active

# Count repeated PDBs
pdb_occurrences = {}
for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in col:
        if cell.value and isinstance(cell.value, str):
            pdb_occurrences[cell.value] = pdb_occurrences.get(cell.value, 0) + 1


def generate_pastel_color():
    r = random.randint(150, 255)
    g = random.randint(150, 255)
    b = random.randint(150, 255)
    return f"{r:02X}{g:02X}{b:02X}"

# Assign colors to repeated PDBs
color_map = {pdb: generate_pastel_color() for pdb, count in pdb_occurrences.items() if count > 1}


for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in col:
        if cell.value in color_map:
            cell.fill = PatternFill(start_color=color_map[cell.value], fill_type="solid")

# Save the colored Excel file
wb.save(filtered_output_file)

# Count PDB occurrences and lysine groups
df2 = pd.read_excel(filtered_output_file, header=None)

lys_columns = df2.iloc[0].dropna().tolist()
df2 = df2.drop(index=[0, 1]).reset_index(drop=True)
df2.columns = lys_columns


melted_df = df2.melt(value_name="Value", var_name="Lys").dropna()
melted_df["Value"] = melted_df["Value"].astype(str).str.strip()

summary_df = (
    melted_df.groupby("Value")
    .agg(Count=("Value", "size"), Lys=("Lys", lambda x: ", ".join(sorted(set(x)))))
    .reset_index()
)

# Save final summary
summary_output_file = "modified_pdbs/Lys_counts_RMSD.xlsx"
summary_df.to_excel(summary_output_file, index=False)

print(f"✅ All steps completed. Final output saved to: {summary_output_file}")
